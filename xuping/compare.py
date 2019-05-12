#!/usr/bin/python
# -*- coding: UTF-8 -*- 

import openpyxl
from xuping import config
from openpyxl.styles import PatternFill
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr

def compare(file):
    wb = openpyxl.load_workbook(file)
    ws1 = wb.get_sheet_by_name('订单表数据')
    ws2 = wb.get_sheet_by_name('第三方数据')
    dict1 = config.pull(ws1)
    dict2 = config.pull(ws2)

    dict1['故意错的'] = 123
    dict2['故意错的'] = 12345

    ws = wb.create_sheet()
    ws.title = '数据校验结果'
    ws.append(['订单号', '订单表价格', '第三方价格', '校验结果'])

    for k in dict1.keys():
        if dict1[k] == dict2[k]:
            ws.append([k, dict1[k], dict2[k], '校验成功'])
        else:
            ws.append([k, dict1[k], dict2[k], '校验失败'])
            ws.cell(ws.max_row,4).fill = PatternFill(fill_type='solid', fgColor='FF0000')

    config.sheet_layout(ws)
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 30

    wb.save(file)

    # 发邮件
    my_sender = 'zhangt@shinemo.com'  # 发件人邮箱账号，为了后面易于维护，所以写成了变量
    my_user = 'zhangt@shinemo.com'  # 收件人邮箱账号，为了后面易于维护，所以写成了变量
    try:
        msg = MIMEText('填写邮件内容', 'plain', 'utf-8')
        msg['From'] = formataddr(["发件人邮箱昵称", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        msg['To'] = formataddr(["收件人邮箱昵称", my_user])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
        msg['Subject'] = "主题"  # 邮件的主题，也可以说是标题
        server = smtplib.SMTP("smtp.exmail.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
        server.login(my_sender, "发件人邮箱密码")  # 括号中对应的是发件人邮箱账号、邮箱密码
        server.sendmail(my_sender, [my_user, ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
        server.quit()  # 这句是关闭连接的意思

    except:
        print('123')# 如果try中的语句没有执行，则会执行下面的ret=False

