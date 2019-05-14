#!/usr/bin/python
# -*- coding: UTF-8 -*- 

from openpyxl.styles import Font, Border, Side, Alignment
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr

'''数据库连接配置'''
DB_CONFIG = {
    'host': '10.0.10.42',
    'port': 3306,
    'user': 'root',
    'passwd': 'shinemo123',
    'db': 'buy'
}

'''订单表数据SQL查询语句'''
SQL_order_price = 'select order_id,price,update_user_name,gmt_pay from buy_order_new where biz_type = 3 ' \
                  'and order_status = 12 order by id desc'

'''第三方数据SQL查询语句'''
SQL_third_data = 'select order_id,price,update_user_name,gmt_pay from buy_order_new where biz_type = 3 ' \
                  'and order_status = 12'

'''Excel表格格式优化'''
def sheet_layout(ws):
    # 设置单元格字体为11号黑体
    font = Font(name='黑体', size=11)
    # 设置单元格格式为水平居中，垂直居中
    alignment = Alignment(horizontal='center', vertical='center')
    # 设置单元格边框颜色为黑色
    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    for column in ws.columns:
        for cell in column:
            cell.alignment = alignment
            cell.font = font
            cell.border = border

'''提取Excel数据到字典'''
def pull(ws):
    dict = {}
    for x in range(2, ws.max_row + 1):
        dict[ws.cell(x, 1).value] = ws.cell(x, 2).value
    return dict

'''发送邮件'''
def mail(my_user, contant, title):
    my_sender = 'zhangt@shinemo.com'  # 发件人邮箱账号，为了后面易于维护，所以写成了变量
    try:
        msg = MIMEText(contant, 'plain', 'utf-8')
        msg['From'] = formataddr(["系统通知", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        msg['Subject'] = title  # 邮件的主题，也可以说是标题
        server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
        server.login(my_sender, "12345qWE")  # 括号中对应的是发件人邮箱账号、邮箱密码
        server.sendmail(my_sender, my_user, msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
        server.quit()  # 这句是关闭连接的意思

    except:
        print('邮件发送失败！')