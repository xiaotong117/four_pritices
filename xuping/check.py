#!/usr/bin/python
# -*- coding: UTF-8 -*- 

import openpyxl
import logging
import difflib
from xuping import tools,config
from openpyxl.styles import PatternFill

# def compare(file):
#     wb = openpyxl.load_workbook(file)
#     ws1 = wb.get_sheet_by_name('测试数据')
#     ws2 = wb.get_sheet_by_name('预期数据')
#     dict1 = tools.pull(ws1)
#     dict2 = tools.pull(ws2)
#     mail_flag = 0
#     contant = []
#
#     dict1['99999999999999999'] = 123
#     dict2['99999999999999999'] = 12345
#
#     ws = wb.create_sheet()
#     ws.title = '数据校验结果'
#     ws.append(['订单号', '订单表价格', '第三方价格', '校验结果'])
#
#     for k in list(dict1.keys()):
#         if k in dict2:
#             if dict1[k] == dict2[k]:
#                 ws.append([k, dict1[k], dict2[k], '校验成功'])
#                 del dict1[k]
#                 del dict2[k]
#             else:
#                 ws.append([k, dict1[k], dict2[k], '校验失败'])
#                 ws.cell(ws.max_row,4).fill = PatternFill(fill_type='solid', fgColor='FF0000')
#                 mail_flag = 1
#                 contant.append(k)
#                 del dict1[k]
#                 del dict2[k]
#         else:
#             ws.append([k, dict1[k], 'Null', '校验失败'])
#             ws.cell(ws.max_row, 4).fill = PatternFill(fill_type='solid', fgColor='FF0000')
#             mail_flag = 1
#             contant.append(k)
#     if dict2:
#         for k in dict2.keys():
#             ws.append([k, 'Null', dict2[k], '校验失败'])
#             ws.cell(ws.max_row, 4).fill = PatternFill(fill_type='solid', fgColor='FF0000')
#             mail_flag = 1
#             contant.append(k)
#
#     tools.sheet_layout(ws)
#     for col in ['A', 'B', 'C', 'D']:
#         ws.column_dimensions[col].width = 30
#
#     wb.save(file)
#
#     if mail_flag:
#         text = ''
#         for x in contant:
#             text += x + '\n'
#         message = '下列订单对账出错：\n'+ text + '\n详情请查看业务金额核对脚本结果！'
#         tools.mail('zhangt@shinemo.com', message, '对账出错了！')


class check_status(object):
    def __init__(self, order_list):
        self.order_list = order_list

    def check_base(order_list):
        false = {}
        db_buy = tools.db_connect(config.DB_CONFIG_BUY)
        db_tc = tools.db_connect(config.DB_CONFIG_TC)
        db_welfare = tools.db_connect(config.DB_CONFIG_WELFARE)
        b_data = tools.pull_data(db_buy, [config.SQL_buy_order_new1], tuple(order_list))
        t_data = tools.pull_data(db_tc, [config.SQL_user_frezen_detail], tuple(order_list))
        w_data = tools.pull_data(db_welfare, [config.SQL_account_freeze], tuple(order_list))
        tools.db_disconnect(db_buy)
        tools.db_disconnect(db_tc)
        tools.db_disconnect(db_welfare)

        for b in b_data[0]:
            try:
                if b['order_status'] in [-1,0]:
                    if b['order_type'] == 2 and b['third_id'] != '' and b['price'] != 0:
                        pass
                    else:
                        false[b['order_id']] = 'buy_order_new表校验失败！'
                        continue
                else:
                    false[b['order_id']] = '订单状态错误！'
                    continue
            except:
                false[b['order_id']] = '订单状态错误！'

            if b['order_status'] == -1:
                pass
            else:
                if b['source_app'] == 0:
                    try:
                        t = difflib.get_close_matches(b['order_id'], t_data[0])
                        if t['biz_type'] == 3 and t['money'] == b['price'] and t['frozen_status'] == 1:
                            pass
                        else:
                            false[b['order_id']] = 'user_frezen_detail表校验失败！'
                            continue
                    except:
                        false[b['order_id']] = 'user_frezen_detail表校验失败！'

                elif b['source_app'] == 8:
                    try:
                        w = difflib.get_close_matches(b['order_id'], w_data[0])
                        if w['amount'] == b['price']*1.5 and w['freeze_status'] == 1:
                            pass
                        else:
                            false[b['order_id']] = 'account_freeze表校验失败！'
                            continue
                    except:
                        false[b['order_id']] = 'account_freeze表校验失败！'

                elif b['source_app'] == 13:
                    try:
                        w = difflib.get_close_matches(b['order_id'], w_data[0])
                        if w['amount'] == b['price']*2 and w['freeze_status'] == 1:
                            pass
                        else:
                            false[b['order_id']] = 'account_freeze表校验失败！'
                            continue
                    except:
                        false[b['order_id']] = 'account_freeze表校验失败！'

                else:
                    false[b['order_id']] = '订单类型错误！'
                    continue
        return false

    def check_20001(order_list):
        false = {}
        db_buy = tools.db_connect(config.DB_CONFIG_BUY)
        data = tools.pull_data(db_buy, [config.SQL_car_order1], tuple(order_list))
        tools.db_disconnect(db_buy)

        for x in data[0]:
            order_list.remove(x['order_id'])
            try:
                if x['biz_type'] == 3 and all(value != '' for value in x.values()):
                    pass
                else:
                    false[x['order_id']] = 'car_order表校验失败！'
                    continue
            except:
                false[x['order_id']] = 'car_order表校验失败！'
        if order_list:
            for x in order_list:
                false[x] = 'car_order表校验失败！'
        return false

    def check_20003(order_list):
        false = {}
        db_buy = tools.db_connect(config.DB_CONFIG_BUY)
        data = tools.pull_data(db_buy, [config.SQL_car_order2], tuple(order_list))
        tools.db_disconnect(db_buy)

        for x in data[0]:
            order_list.remove(x['order_id'])
            try:
                if x['biz_type'] == 3 and all(value != '' for value in x.values()):
                    pass
                else:
                    false[x['order_id']] = 'car_order表校验失败！'
                    continue
            except:
                false[x['order_id']] = 'car_order表校验失败！'
        if order_list:
            for x in order_list:
                false[x] = 'car_order表校验失败！'
        return false

    def check_20007(order_list):
        false = {}
        db_buy = tools.db_connect(config.DB_CONFIG_BUY)
        data = tools.pull_data(db_buy, [config.SQL_car_order3], tuple(order_list))
        tools.db_disconnect(db_buy)

        for x in data[0]:
            order_list.remove(x['order_id'])
            try:
                if x['biz_type'] == 3 and all(value != '' for value in x.values()):
                    pass
                else:
                    false[x['order_id']] = 'car_order表校验失败！'
                    continue
            except:
                false[x['order_id']] = 'car_order表校验失败！'
        if order_list:
            for x in order_list:
                false[x] = 'car_order表校验失败！'
        return false

    def check_20009(order_list):
        false = {}
        db_buy = tools.db_connect(config.DB_CONFIG_BUY)
        data = tools.pull_data(db_buy, [config.SQL_car_order4], tuple(order_list))
        tools.db_disconnect(db_buy)

        for x in data[0]:
            order_list.remove(x['order_id'])
            try:
                if x['biz_type'] == 3 and all(value != '' for value in x.values()):
                    pass
                else:
                    false[x['order_id']] = 'car_order表校验失败！'
                    continue
            except:
                false[x['order_id']] = 'car_order表校验失败！'
        if order_list:
            for x in order_list:
                false[x] = 'car_order表校验失败！'
        return false

    def check_20017(order_list):
        false = {}
        db_buy = tools.db_connect(config.DB_CONFIG_BUY)
        db_tc = tools.db_connect(config.DB_CONFIG_TC)
        db_welfare = tools.db_connect(config.DB_CONFIG_WELFARE)
        b_data = tools.pull_data(db_buy, [config.SQL_buy_order_new2, config.SQL_car_order5], tuple(order_list))
        t_data = tools.pull_data(db_tc, [config.SQL_user_frezen_detail, config.SQL_pay_detail], tuple(order_list))
        w_data = tools.pull_data(db_welfare, [config.SQL_account_freeze, config.SQL_welfare_turnover], tuple(order_list))
        tools.db_disconnect(db_buy)
        tools.db_disconnect(db_tc)
        tools.db_disconnect(db_welfare)

        for b in b_data[0]:
            try:
                if b['source_app'] == 0:
                    if b['order_type'] == 2 and  b['order_status'] == 12 and b['third_id'] != "" and \
                            b['gmt_pay'] != "" and b['gmt_settlement_time'] != "" and b['price'] != 0 and \
                            b['should_price'] != 0:
                        pass
                    else:
                        false[b['order_id']] = 'buy_order表校验失败！'
                        continue
                if b['source_app'] == 8 or b['source_app'] == 13:
                    if b['order_type'] == 2 and b['order_status'] == 12 and b['third_id'] != "" and \
                            b['gmt_pay'] != "" and b['gmt_settlement_time'] != "" and b['price'] != 0 and \
                            b['integral_price'] != 0:
                        pass
                    else:
                        false[b['order_id']] = 'buy_order表校验失败！'
                        continue

                b1 = difflib.get_close_matches(b['order_id'], b_data[1])
                if b1['biz_type'] == 3 and all(value != '' for value in b1.values()):
                    pass
                else:
                    false[b['order_id']] = 'car_order表校验失败！'
                    continue
            except:
                false[b['order_id']] = '订单状态错误！'

            if b['source_app'] == 0:
                try:
                    t = difflib.get_close_matches(b['order_id'], t_data[0])
                    t1 = difflib.get_close_matches(b['order_id'], t_data[1])
                    if t['biz_type'] == 3 and t['money'] !='' and t['frozen_status'] == 2 and \
                            t1['biz_type'] == 3 and t1['price'] == b['should_price'] and t1['pay_status'] == 6:
                        pass
                    else:
                        false[b['order_id']] = 'user_frezen_detail/pay_detail表校验失败！'
                        continue
                except:
                    false[b['order_id']] = 'user_frezen_detail/pay_detail表校验失败！'

            elif b['source_app'] == 8 or b['source_app'] == 13:
                try:
                    w = difflib.get_close_matches(b['order_id'], w_data[0])
                    w1 = difflib.get_close_matches(b['order_id'], w_data[1])
                    if w['amount'] != '' and w['freeze_status'] == 0 and w1['account_type'] == 0 and \
                            w1['turnover_type'] == 2 and w1['turnover_sub_type'] == 3 and w1['remark']\
                            == '滴滴出行' and w1['amount'] == b['should_price'] and w1['turnover_status'] == 3:
                        pass
                    else:
                        false[b['order_id']] = 'account_freeze/welfare_turnover表校验失败！'
                        continue
                except:
                    false[b['order_id']] = 'account_freeze/welfare_turnover表校验失败！'

            else:
                false[b['order_id']] = '订单类型错误！'
                continue
        return false

    def check_20019(order_list):
        false = {}
        db_buy = tools.db_connect(config.DB_CONFIG_BUY)
        db_tc = tools.db_connect(config.DB_CONFIG_TC)
        db_welfare = tools.db_connect(config.DB_CONFIG_WELFARE)
        b_data = tools.pull_data(db_buy, [config.SQL_buy_order_new2, config.SQL_car_order1, config.SQL_car_order3],
                                 tuple(order_list))
        t_data = tools.pull_data(db_tc, [config.SQL_user_frezen_detail, config.SQL_pay_detail], tuple(order_list))
        w_data = tools.pull_data(db_welfare, [config.SQL_account_freeze, config.SQL_welfare_turnover],
                                 tuple(order_list))
        tools.db_disconnect(db_buy)
        tools.db_disconnect(db_tc)
        tools.db_disconnect(db_welfare)

        for b in b_data[0]:
            if b['price'] == 0:
                if b['order_type'] == 2 and b['order_status'] == 18 and b['third_id'] != '':
                    pass
                else:
                    false[b['order_id']] = 'buy_order_new表校验失败！'
                    continue

                try:
                    b1 = difflib.get_close_matches(b['order_id'], b_data[1])
                    if b[0] == 3 and all(x != '' for x in b):
                        pass
                    else:
                        false[b['order_id']] = 'car_order表校验失败！'
                        continue
                except:
                    false[b['order_id']] = 'car_order表校验失败！'

                if b[8] == 0:
                    try:
                        t_data = tools.pull_data(db_tc, [config.SQL_user_frezen_detail], order)
                        c = t_data[0]
                        if c[0] == 3 and c[1] != '' and c[2] == 2:
                            pass
                        else:
                            false[b['order_id']] = 'user_frezen_detail表校验失败！'
                            continue
                    except:
                        false[b['order_id']] = 'user_frezen_detail表校验失败！'

                if b[8] == 8 or b[8] == 13:
                    try:
                        t_data = tools.pull_data(db_welfare, [config.SQL_account_freeze], order)
                        c = t_data[0]
                        if c[0] != '' and c[1] == 0:
                            pass
                        else:
                            false[b['order_id']] = 'account_freeze表校验失败！'
                            continue
                    except:
                        false[b['order_id']] = 'account_freeze表校验失败！'

            elif b['price'] != 0:
                if b['order_type'] == 2 and b['order_status'] == 18 and b['third_id'] != '' and b['gmt_pay'] \
                        != '' and b['gmt_settlement_time'] != '' and b['price'] != 0:
                    pass
                else:
                    false[b['order_id']] = 'buy_order_new表校验失败！'
                    continue

                try:
                    b2 = difflib.get_close_matches(b['order_id'], b_data[2])
                    if b[0] == 3 and all(x != '' for x in b):
                        pass
                    else:
                        false[b['order_id']] = 'car_order表校验失败！'
                        continue
                except:
                    false[b['order_id']] = 'car_order表校验失败！'

                if b[8] == 0:
                    try:
                        t_data = tools.pull_data(db_tc, [config.SQL_user_frezen_detail, config.SQL_pay_detail], order)
                        c = t_data[0]
                        d = t_data[1]
                        if c[0] == 3 and c[1] != '' and c[2] == 2 and d[0] == 3 and d[1] == a[6] and d[2] == 6:
                            pass
                        else:
                            false[b['order_id']] = 'user_frezen_detail/pay_detail表校验失败！'
                            continue
                    except:
                        false[b['order_id']] = 'user_frezen_detail/pay_detail表校验失败！'

                elif b[8] == 8 or b[8] == 13:
                    try:
                        t_data = tools.pull_data(db_welfare, [config.SQL_account_freeze, config.SQL_welfare_turnover], order)
                        c = t_data[0]
                        d = t_data[1]
                        if c[0] != '' and c[1] == 0 and d[0] == 0 and d[1] == 2 and d[2] == 3 and d[3] == '滴滴出行' and d[4] == \
                                a[6] and d[6] == 3:
                            pass
                        else:
                            false[b['order_id']] = 'account_freeze/welfare_turnover表校验失败！'
                            continue
                    except:
                        false[b['order_id']] = 'account_freeze/welfare_turnover表校验失败！'
        return false


    def check_20021(order_list):
        false = {}
        db_buy = tools.db_connect(config.DB_CONFIG_BUY)
        db_tc = tools.db_connect(config.DB_CONFIG_TC)
        db_welfare = tools.db_connect(config.DB_CONFIG_WELFARE)
        for order in order_list:
            b_data = tools.pull_data(db_buy, [config.SQL_buy_order_new1, config.SQL_car_order1], order)
            a = b_data[0]
            if a[0] == 2 and a[1] == 18 and a[2] != '' and a[3] == 0:
                pass
            else:
                false[order] = 'buy_order_new表校验失败！'
                continue

            try:
                b = b_data[1]
                if b[0] == 3 and all(x != '' for x in b):
                    pass
                else:
                    false[order] = 'car_order表校验失败！'
                    continue
            except:
                'car_order表校验失败！'

            if a[4] == 0:
                try:
                    t_data = tools.pull_data(db_tc, [config.SQL_user_frezen_detail], order)
                    c = t_data[0]
                    if c[0] == 3 and c[1] != '' and c[2] == 2:
                        pass
                    else:
                        false[order] = 'user_frezen_detail表校验失败！'
                        continue
                except:
                    false[order] = 'user_frezen_detail表校验失败！'

            if a[4] == 8 or a[4] == 13:
                try:
                    t_data = tools.pull_data(db_welfare, [config.SQL_account_freeze], order)
                    c = t_data[0]
                    if c[0] != '' and c[1] == 0:
                        pass
                    else:
                        false[order] = 'account_freeze表校验失败！'
                        continue
                except:
                    false[order] = 'account_freeze表校验失败！'
        tools.db_disconnect(db_buy)
        tools.db_disconnect(db_tc)
        tools.db_disconnect(db_welfare)
        return false






