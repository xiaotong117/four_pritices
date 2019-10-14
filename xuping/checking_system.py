#!/usr/bin/python
# -*- coding: UTF-8 -*- 

import openpyxl
import time
import logging
import win32api
from xuping.db_interaction import get_db_price
from xuping.check import check_status
from xuping import tools

if __name__ == "__main__":
    nowdate = time.strftime("%Y-%m-%d", time.localtime())
    file = '业务金额核对' + nowdate +'.xlsx'
    break_flag = 1
    print('************************************')
    print('*********滴滴订单数据校验系统********')
    print('************************************')
    while break_flag:
        time.sleep(1)
        print('\n请输入操作代码：\n1、预加载数据；2、数据校验；3、查看结果；4、退出系统')
        a = input()
        if a == '1':
            try:
                start = time.time()
                get_db_price(file)
                end = time.time()
                print(end - start)
                print('数据加载成功！')
            except:
                logging.error("哎呀，出错了！", exc_info=True)
        if a == '2':
            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.create_sheet()
                ws.title = '数据校验结果' + time.strftime("%H-%M-%S", time.localtime())
                ws.append(['订单子状态','订单号', '出错字段'])

                # for x in ['20001', '20003', '20007', '20009']:
                for x in ['20001']:
                    ws1 = wb.get_sheet_by_name('滴滴订单')
                    listx = [ws1.cell(y, 1).value for y in range(2, ws1.max_row + 1) if ws1.cell(y, 2).value == int(x)]
                    print('校验状态%s数据%s条。'%(x, len(listx)))
                    false = check_status.check_base(listx)
                    # c = {'20001':check_status.check_20001(listx), '20003':check_status.check_20003(listx), '20007':check_status.check_20007(listx), '20009':check_status.check_20009(listx)}[x]
                    c = {'20001':check_status.check_20001(listx)}[x]

                    false.update(c)
                    print('其中%s条数据异常。\n'%(len(false)))
                    for k in list(false.keys()):
                        ws.append([x, k, false[k]])

                # for x in ('20017', '20019', '20021'):
                #     ws1 = wb.get_sheet_by_name('滴滴订单')
                #     listx = [ws1.cell(y, 1).value for y in range(2, ws1.max_row + 1) if ws1.cell(y, 2).value == int(x)]
                #     print('校验状态%s数据%s条。'%(x, len(listx)))
                #     false = {'20017':check_status.check_20017(listx), '20019':check_status.check_20019(listx), '20021':check_status.check_20021(listx)}[x]
                #     print('其中%s条数据异常。\n'%(len(false)))
                #     for k in list(false.keys()):
                #         ws.append([x, k, false[k]])


                # ws1 = wb.get_sheet_by_name('滴滴订单')
                # xxx = '20001'
                #
                # listx = [ws1.cell(x, 1).value for x in range(2, ws1.max_row + 1) if ws1.cell(x, 2).value == int(xxx)]
                # # listx = []
                # # for x in listy:
                # #     listx.append(x)
                # print('校验状态20001数据%s条。' % (len(listx)))
                # false = check_status.check_20001(listx)
                # print('其中%s条数据异常。\n' % (len(false)))
                # for k in list(false.keys()):
                #     ws.append(['20001', k, false[k]])

                tools.sheet_layout(ws)
                for col in ['A', 'B', 'C']:
                    ws.column_dimensions[col].width = 40

                wb.save(file)
            except:
                logging.error("哎呀，出错了！", exc_info=True)
        if a == '3':
            try:
                win32api.ShellExecute(0, 'open', file, '', '', 1)
            except:
                logging.error("哎呀，出错了！", exc_info=True)
        if a == '4':
            break_flag = 0
